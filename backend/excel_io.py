"""Импорт и экспорт плана в Excel.

В файле предшественники записаны именами, в базе — id. Разбор идёт в два
прохода: порядок строк произволен, и ссылка вперёд по файлу иначе теряется.
"""

import io
from datetime import date, datetime

from openpyxl import Workbook, load_workbook

COLUMNS = ["Задача", "Описание", "Исполнитель", "Длительность", "Предшественники"]
CONSTRAINT_COLUMN = "Не раньше"
FIRST_DATA_ROW = 2


def _split_names(cell) -> list[str]:
    if cell in (None, ""):
        return []
    return [part.strip() for part in str(cell).split(",") if part.strip()]


def parse_excel(data: bytes):
    """Разобрать файл. Возвращает (задачи, ошибки).

    Ошибки собираются все сразу и указывают на строку и колонку. Если ошибка
    хотя бы одна, задачи не возвращаются: частичная загрузка не ломает расчёты,
    а даёт молча неверный ответ.
    """
    try:
        workbook = load_workbook(io.BytesIO(data))
    except Exception:
        return [], [{"row": 0, "column": "Файл",
                     "message": "не удалось прочитать файл: ожидается книга Excel (.xlsx)"}]
    sheet = workbook.active
    header = next(sheet.iter_rows(max_row=1, values_only=True), ()) or ()
    header = [str(cell).strip() if cell is not None else "" for cell in header]
    if header[:len(COLUMNS)] != COLUMNS:
        return [], [{"row": 1, "column": "Шапка",
                     "message": f"ожидаются колонки: {', '.join(COLUMNS)}"}]
    has_constraint_column = len(header) > len(COLUMNS) and header[len(COLUMNS)] == CONSTRAINT_COLUMN

    rows = list(sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True))

    errors = []
    names = {}
    tasks = []

    # Первый проход: раздать id, заполнить словарь имён. Связи не трогаем.
    for offset, row in enumerate(rows):
        row_number = FIRST_DATA_ROW + offset
        name = (row[0] or "").strip() if row and row[0] else ""
        if not name and not any(cell is not None and cell != "" for cell in (row or [])):
            continue
        if not name:
            errors.append({"row": row_number, "column": "Задача",
                           "message": "пустое имя задачи"})
            continue
        if "," in name:
            errors.append({"row": row_number, "column": "Задача",
                           "message": f"имя «{name}» содержит запятую: она разделяет предшественников"})
            continue
        if name in names:
            errors.append({"row": row_number, "column": "Задача",
                           "message": f"задача «{name}» встречается второй раз"})
            continue

        raw = row[3]
        if raw is None:
            errors.append({"row": row_number, "column": "Длительность",
                           "message": "длительность не указана"})
            continue
        try:
            if isinstance(raw, bool) or float(raw) != int(float(raw)):
                raise ValueError
            duration = int(float(raw))
            if duration < 0:
                errors.append({"row": row_number, "column": "Длительность",
                               "message": f"«{raw}» — длительность не может быть отрицательной"})
                continue
        except (TypeError, ValueError):
            errors.append({"row": row_number, "column": "Длительность",
                           "message": f"«{raw}» не целое число дней"})
            continue

        # Колонка с ограничением опциональна: файлы, выгруженные до её появления,
        # остаются в силе. Читаем шестую ячейку, только если шапка её подтвердила —
        # иначе чужая колонка (например, «Комментарий») читалась бы как ограничение.
        # Ячейка может прийти датой (Excel) или строкой (правка руками).
        raw_constraint = row[5] if has_constraint_column and row and len(row) > 5 else None
        not_earlier_than = None
        if raw_constraint not in (None, ""):
            if isinstance(raw_constraint, datetime):
                not_earlier_than = raw_constraint.date().isoformat()
            elif isinstance(raw_constraint, date):
                not_earlier_than = raw_constraint.isoformat()
            else:
                try:
                    date.fromisoformat(str(raw_constraint).strip())
                    not_earlier_than = str(raw_constraint).strip()
                except ValueError:
                    errors.append({"row": row_number, "column": CONSTRAINT_COLUMN,
                                   "message": f"«{raw_constraint}» не дата в формате ГГГГ-ММ-ДД"})
                    continue

        task_id = len(tasks) + 1
        names[name] = task_id
        tasks.append({"id": task_id, "name": name, "description": row[1] or "",
                      "assignee": row[2] or "", "duration": duration,
                      "predecessors": [], "not_earlier_than": not_earlier_than,
                      "_row": row_number, "_predecessor_names": _split_names(row[4])})

    # Второй проход: те же строки, теперь любое имя находится.
    for task in tasks:
        for predecessor_name in task["_predecessor_names"]:
            if predecessor_name not in names:
                errors.append({"row": task["_row"], "column": "Предшественники",
                               "message": f"задача «{predecessor_name}» не найдена"})
                continue
            predecessor_id = names[predecessor_name]
            if predecessor_id not in task["predecessors"]:
                task["predecessors"].append(predecessor_id)

    if not errors:
        errors.extend(_check_cycles(tasks))

    for task in tasks:
        del task["_row"]
        del task["_predecessor_names"]

    return ([], errors) if errors else (tasks, [])


def _check_cycles(tasks) -> list[dict]:
    """Тот же обход, что в recalculate, но без арифметики дат.

    Цикл ловится здесь, а не при пересчёте: тут известны имена и номера строк,
    а в recalculate только id — пользователь получил бы «Цикл: [1, 2, 3]».
    """
    indegree = {t["id"]: len(t["predecessors"]) for t in tasks}
    successors = {t["id"]: [] for t in tasks}
    for task in tasks:
        for p in task["predecessors"]:
            successors[p].append(task["id"])

    ready = [tid for tid, deg in indegree.items() if deg == 0]
    seen = 0
    while ready:
        tid = ready.pop()
        seen += 1
        for s in successors[tid]:
            indegree[s] -= 1
            if indegree[s] == 0:
                ready.append(s)

    if seen == len(tasks):
        return []

    stuck = [t for t in tasks if indegree[t["id"]] > 0]
    listed = ", ".join(f"«{t['name']}»" for t in stuck)
    return [{"row": min(t["_row"] for t in stuck), "column": "Предшественники",
             "message": f"циклическая зависимость между задачами: {listed}"}]


def export_excel(plan) -> bytes:
    """Выгрузить план. Предшественники пишутся именами: файл должен грузиться обратно.

    Имена задач не должны повторяться: список предшественников через запятую
    иначе станет неоднозначным, и файл нельзя будет загрузить обратно
    (parse_excel отклоняет повторяющиеся имена).
    """
    names = {task["id"]: task["name"] for task in plan["tasks"]}
    seen = set()
    for name in names.values():
        if name in seen:
            raise ValueError(f"повторяющееся имя задачи «{name}»: такой файл не загрузится обратно")
        seen.add(name)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "План"
    sheet.append(COLUMNS + [CONSTRAINT_COLUMN])
    for task in plan["tasks"]:
        sheet.append([
            task["name"], task["description"], task["assignee"], task["duration"],
            ", ".join(names[p] for p in task["predecessors"]),
            task["not_earlier_than"] or "",
        ])
    for column, width in zip("ABCDEF", (30, 40, 18, 14, 40, 16)):
        sheet.column_dimensions[column].width = width
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
