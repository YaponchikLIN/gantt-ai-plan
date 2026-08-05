from datetime import date
import io
import json
import os

import httpx

import agent_loop
import excel_io
import plan_mcp_server as srv


def set_tasks(tasks, project_start=date(2026, 9, 1)):
    """Положить план в хранилище напрямую, минуя инструменты."""
    srv.STATE["tasks"] = {t["id"]: t for t in tasks}
    srv.STATE["project_start"] = project_start
    srv.STATE["next_id"] = max((t["id"] for t in tasks), default=0) + 1


def task(id, name, duration, predecessors=(), assignee="", not_earlier_than=None):
    return {"id": id, "name": name, "description": "", "assignee": assignee,
            "duration": duration, "predecessors": list(predecessors),
            "not_earlier_than": not_earlier_than}


def by_name(plan):
    return {t["name"]: t for t in plan["tasks"]}


def test_chain_of_three_gets_six_computed_dates():
    set_tasks([
        task(1, "Проектирование", 5),
        task(2, "Разработка", 10, [1]),
        task(3, "Тестирование", 4, [2]),
    ])
    t = by_name(srv.recalculate())
    assert (t["Проектирование"]["start"], t["Проектирование"]["finish"]) == ("2026-09-01", "2026-09-06")
    assert (t["Разработка"]["start"], t["Разработка"]["finish"]) == ("2026-09-06", "2026-09-16")
    assert (t["Тестирование"]["start"], t["Тестирование"]["finish"]) == ("2026-09-16", "2026-09-20")


def test_constraint_pushes_task_right():
    set_tasks([
        task(1, "Проектирование", 5),
        task(2, "Разработка", 10, [1], not_earlier_than=date(2026, 9, 10)),
    ])
    t = by_name(srv.recalculate())
    assert t["Разработка"]["start"] == "2026-09-10"


def test_cycle_raises_with_ids():
    set_tasks([
        task(1, "А", 1, [3]),
        task(2, "Б", 1, [1]),
        task(3, "В", 1, [2]),
    ])
    try:
        srv.recalculate()
    except ValueError as e:
        assert "1" in str(e) and "2" in str(e) and "3" in str(e)
    else:
        raise AssertionError("цикл должен был вызвать ValueError")


def test_missing_predecessor_raises():
    set_tasks([task(1, "А", 3, [99])])
    try:
        srv.recalculate()
    except ValueError as e:
        assert "99" in str(e)
    else:
        raise AssertionError("ссылка на несуществующую задачу должна была вызвать ValueError")


def branching_plan():
    """Проектирование → (Разработка | Документация) → Тестирование."""
    return [
        task(1, "Проектирование", 5),
        task(2, "Разработка", 4, [1]),
        task(3, "Документация", 7, [1]),
        task(4, "Тестирование", 4, [2, 3]),
    ]


def test_slack_and_critical_path():
    set_tasks(branching_plan())
    plan = srv.recalculate()
    t = by_name(plan)
    assert t["Тестирование"]["start"] == "2026-09-13"
    assert t["Разработка"]["slack"] == 3
    assert t["Разработка"]["on_critical_path"] is False
    assert t["Документация"]["slack"] == 0
    assert t["Документация"]["on_critical_path"] is True
    assert plan["project_finish"] == "2026-09-17"


def test_delay_of_five_days_with_slack_three_moves_end_by_two():
    tasks = branching_plan()
    tasks[1]["duration"] = 4 + 5  # Разработка опаздывает на 5 дней
    set_tasks(tasks)
    plan = srv.recalculate()
    t = by_name(plan)
    assert plan["project_finish"] == "2026-09-19"  # было 17-е, уехало на 2
    assert t["Разработка"]["on_critical_path"] is True
    assert t["Документация"]["on_critical_path"] is False


def test_get_plan_returns_computed_dates():
    set_tasks(branching_plan())
    plan = srv.get_plan()
    assert len(plan["tasks"]) == 4
    assert by_name(plan)["Проектирование"]["start"] == "2026-09-01"


def test_load_plan_replaces_everything():
    set_tasks(branching_plan())
    plan = srv._load_plan(
        tasks=[{"id": 1, "name": "Одна", "description": "", "assignee": "",
                "duration": 3, "predecessors": [], "not_earlier_than": None}],
        project_start="2026-10-01",
    )
    assert [t["name"] for t in plan["tasks"]] == ["Одна"]
    assert plan["project_start"] == "2026-10-01"
    assert srv.STATE["next_id"] == 2


def test_transaction_rolls_back_on_failure():
    set_tasks(branching_plan())
    before = srv.recalculate()
    try:
        with srv._transaction():
            srv.STATE["tasks"][2]["predecessors"].append(4)  # замыкает цикл
    except ValueError:
        pass
    else:
        raise AssertionError("цикл должен был вызвать ValueError")
    assert srv.recalculate() == before


def test_load_plan_rolls_back_on_bad_data():
    set_tasks(branching_plan())
    before = srv.recalculate()
    try:
        srv._load_plan(
            tasks=[{"id": 1, "name": "Одна", "duration": 3, "predecessors": [99]}],
            project_start="2026-10-01",
        )
    except ValueError:
        pass
    else:
        raise AssertionError("висящий предшественник должен был вызвать ValueError")
    assert srv.recalculate() == before


def test_load_plan_rejects_duplicate_ids():
    set_tasks(branching_plan())
    before = srv.recalculate()
    try:
        srv._load_plan(
            tasks=[{"id": 1, "name": "Одна", "duration": 3, "predecessors": []},
                   {"id": 1, "name": "Другая", "duration": 2, "predecessors": []}],
            project_start="2026-10-01",
        )
    except ValueError as e:
        assert "1" in str(e)
    else:
        raise AssertionError("повторяющийся id должен был вызвать ValueError")
    assert srv.recalculate() == before


def test_shift_tasks_moves_task_and_its_dependents():
    set_tasks(branching_plan())
    plan = srv.shift_tasks(task_ids=[1], days=2)
    t = by_name(plan)
    assert t["Проектирование"]["start"] == "2026-09-03"
    assert t["Разработка"]["start"] == "2026-09-08"  # наследник уехал сам


def test_reassign_does_not_touch_dates():
    set_tasks(branching_plan())
    before = by_name(srv.recalculate())["Разработка"]["start"]
    plan = srv.reassign_tasks(task_ids=[2], assignee="Петров")
    t = by_name(plan)["Разработка"]
    assert t["assignee"] == "Петров"
    assert t["start"] == before


def test_shift_unknown_task_raises_and_changes_nothing():
    set_tasks(branching_plan())
    before = srv.recalculate()
    try:
        srv.shift_tasks(task_ids=[1, 99], days=5)
    except ValueError as e:
        assert "99" in str(e)
    else:
        raise AssertionError("несуществующая задача должна была вызвать ValueError")
    assert srv.recalculate() == before


def test_set_constraint_applies_not_earlier_than():
    set_tasks(branching_plan())
    plan = srv.set_constraint(task_id=2, not_earlier_than="2026-09-20")
    assert by_name(plan)["Разработка"]["start"] == "2026-09-20"


def test_shift_counts_from_current_start_not_stale_constraint():
    set_tasks(branching_plan())
    srv.set_constraint(task_id=4, not_earlier_than="2026-09-05")  # раньше естественного старта
    assert by_name(srv.recalculate())["Тестирование"]["start"] == "2026-09-13"
    plan = srv.shift_tasks(task_ids=[4], days=7)
    assert by_name(plan)["Тестирование"]["start"] == "2026-09-20"


def test_two_shifts_accumulate():
    set_tasks(branching_plan())
    srv.shift_tasks(task_ids=[1], days=2)
    plan = srv.shift_tasks(task_ids=[1], days=3)
    assert by_name(plan)["Проектирование"]["start"] == "2026-09-06"


def test_add_task_gets_next_id_and_dates():
    set_tasks(branching_plan())
    plan = srv.add_task(name="Приёмка", duration=2, assignee="Сидоров", predecessors=[4])
    t = by_name(plan)["Приёмка"]
    assert t["id"] == 5
    assert t["start"] == "2026-09-17"
    assert t["finish"] == "2026-09-19"


def test_set_duration_moves_dependents():
    set_tasks(branching_plan())
    plan = srv.set_duration(task_id=3, duration=2)  # Документация короче Разработки
    t = by_name(plan)
    assert t["Тестирование"]["start"] == "2026-09-10"
    assert t["Разработка"]["on_critical_path"] is True


def test_add_predecessor_that_closes_cycle_is_rolled_back():
    set_tasks(branching_plan())
    before = srv.recalculate()
    try:
        srv.add_predecessor(task_id=1, predecessor_id=4)  # Тестирование → Проектирование
    except ValueError as e:
        assert "Цикл" in str(e)
    else:
        raise AssertionError("замыкание цикла должно было вызвать ValueError")
    assert srv.recalculate() == before


def test_remove_predecessor_frees_the_task():
    set_tasks(branching_plan())
    plan = srv.remove_predecessor(task_id=3, predecessor_id=1)
    t = by_name(plan)["Документация"]
    assert t["predecessors"] == []
    assert t["start"] == "2026-09-01"


def test_add_predecessor_is_idempotent():
    set_tasks(branching_plan())
    srv.add_predecessor(task_id=4, predecessor_id=2)
    plan = srv.add_predecessor(task_id=4, predecessor_id=2)
    assert by_name(plan)["Тестирование"]["predecessors"] == [2, 3]


def test_add_task_rejects_negative_duration():
    set_tasks(branching_plan())
    before = srv.recalculate()
    try:
        srv.add_task(name="Плохая", duration=-3)
    except ValueError as e:
        assert "Длительность" in str(e)
    else:
        raise AssertionError("отрицательная длительность должна была вызвать ValueError")
    assert srv.recalculate() == before
    assert srv.STATE["next_id"] == 5


def test_add_task_deduplicates_predecessors():
    set_tasks(branching_plan())
    plan = srv.add_task(name="Приёмка", duration=2, predecessors=[4, 4])
    assert by_name(plan)["Приёмка"]["predecessors"] == [4]


def test_remove_absent_predecessor_raises():
    set_tasks(branching_plan())
    before = srv.recalculate()
    try:
        srv.remove_predecessor(task_id=3, predecessor_id=2)
    except ValueError as e:
        assert "2" in str(e)
    else:
        raise AssertionError("отсутствующая связь должна была вызвать ValueError")
    assert srv.recalculate() == before


def test_delete_task_bridges_successors_to_predecessors():
    set_tasks([
        task(1, "Проектирование", 5),
        task(2, "Разработка", 10, [1]),
        task(3, "Тестирование", 4, [2]),
    ])
    result = srv.delete_task(task_id=2)
    t = by_name(result["plan"])
    assert "Разработка" not in t
    assert t["Тестирование"]["predecessors"] == [1]
    assert t["Тестирование"]["start"] == "2026-09-06"  # не уехало на старт проекта
    assert result["rebonded"] == [{"task": "Тестирование",
                                   "inherited": ["Проектирование"],
                                   "predecessors_now": ["Проектирование"]}]


def test_delete_task_without_successors_just_removes_it():
    set_tasks(branching_plan())
    result = srv.delete_task(task_id=4)
    assert "Тестирование" not in by_name(result["plan"])
    assert result["rebonded"] == []
    assert result["plan"]["project_finish"] == "2026-09-13"


def test_delete_task_does_not_duplicate_links():
    set_tasks([
        task(1, "А", 2),
        task(2, "Б", 2, [1]),
        task(3, "В", 2, [1, 2]),  # уже зависит и от А, и от Б
    ])
    result = srv.delete_task(task_id=2)
    assert by_name(result["plan"])["В"]["predecessors"] == [1]


def test_delete_task_reports_only_inherited_links():
    set_tasks([
        task(1, "Аналитика", 3),
        task(2, "Разработка", 5, [1]),
        task(3, "Закупка", 2),
        task(4, "Тестирование", 4, [2, 3]),
    ])
    result = srv.delete_task(task_id=2)
    entry = result["rebonded"][0]
    assert entry["task"] == "Тестирование"
    assert entry["inherited"] == ["Аналитика"]
    assert entry["predecessors_now"] == ["Аналитика", "Закупка"]


def test_delete_unknown_task_raises():
    set_tasks(branching_plan())
    before = srv.recalculate()
    try:
        srv.delete_task(task_id=99)
    except ValueError as e:
        assert "99" in str(e)
    else:
        raise AssertionError("несуществующая задача должна была вызвать ValueError")
    assert srv.recalculate() == before


def test_explain_deadline_when_it_does_not_fit():
    set_tasks(branching_plan())
    info = srv.explain_deadline(task_id=4, required_finish="2026-09-12")
    assert info["fits"] is False
    assert info["days_short"] == 4  # финиш 17-го против границы 13-го (12-е включительно)
    assert info["own_duration"] == 4
    chain = {c["name"]: c for c in info["predecessors_chain"]}
    assert set(chain) == {"Проектирование", "Разработка", "Документация"}
    assert chain["Документация"]["on_critical_path"] is True
    assert chain["Разработка"]["on_critical_path"] is False
    assert [c["name"] for c in info["predecessors_chain"]] == [
        "Документация", "Проектирование", "Разработка"]  # по убыванию длительности


def test_explain_deadline_counts_an_exclusive_deadline_differently():
    set_tasks(branching_plan())
    inclusive = srv.explain_deadline(task_id=4, required_finish="2026-09-12")
    exclusive = srv.explain_deadline(task_id=4, required_finish="2026-09-12", inclusive=False)
    assert inclusive["days_short"] == 4
    assert exclusive["days_short"] == 5
    assert inclusive["required_finish"] == "2026-09-12"  # дата пользователя, не пересчитанная


def test_explain_deadline_when_it_fits():
    set_tasks(branching_plan())
    info = srv.explain_deadline(task_id=4, required_finish="2026-09-30")
    assert info["fits"] is True
    assert info["days_short"] == 0


def test_explain_deadline_changes_nothing():
    set_tasks(branching_plan())
    before = srv.recalculate()
    srv.explain_deadline(task_id=4, required_finish="2026-09-12")
    assert srv.recalculate() == before


def test_explain_deadline_for_task_without_predecessors():
    set_tasks(branching_plan())
    info = srv.explain_deadline(task_id=1, required_finish="2026-09-10")
    assert info["predecessors_chain"] == []
    assert info["fits"] is True
    assert info["on_critical_path"] is True


def make_xlsx(rows):
    """Собрать файл в память: шапка плюс переданные строки."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Задача", "Описание", "Исполнитель", "Длительность", "Предшественники"])
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_import_resolves_forward_reference():
    """Строка ссылается на задачу, которая идёт ниже по файлу."""
    data = make_xlsx([
        ["Тестирование", "", "Петров", 4, "Разработка"],
        ["Разработка", "", "Иванов", 10, "Проектирование"],
        ["Проектирование", "", "Иванов", 5, ""],
    ])
    tasks, errors = excel_io.parse_excel(data)
    assert errors == []
    by_id = {t["id"]: t for t in tasks}
    testing = next(t for t in tasks if t["name"] == "Тестирование")
    assert [by_id[p]["name"] for p in testing["predecessors"]] == ["Разработка"]


def test_import_reports_row_and_column_for_unknown_predecessor():
    data = make_xlsx([
        ["Проектирование", "", "Иванов", 5, ""],
        ["Разработка", "", "Иванов", 10, "Проктирование"],  # опечатка
    ])
    tasks, errors = excel_io.parse_excel(data)
    assert tasks == []
    assert len(errors) == 1
    assert errors[0]["row"] == 3  # шапка занимает первую строку
    assert errors[0]["column"] == "Предшественники"
    assert "Проктирование" in errors[0]["message"]


def test_import_collects_all_errors_at_once():
    data = make_xlsx([
        ["", "", "Иванов", 5, ""],
        ["Разработка", "", "Иванов", "десять", ""],
        ["Тестирование", "", "Петров", 4, "Нет такой"],
    ])
    tasks, errors = excel_io.parse_excel(data)
    assert tasks == []
    assert len(errors) == 3
    assert [e["row"] for e in errors] == [2, 3, 4]


def test_import_catches_cycle_by_names():
    data = make_xlsx([
        ["А", "", "", 1, "В"],
        ["Б", "", "", 1, "А"],
        ["В", "", "", 1, "Б"],
    ])
    tasks, errors = excel_io.parse_excel(data)
    assert tasks == []
    assert len(errors) == 1
    message = errors[0]["message"]
    assert "А" in message and "Б" in message and "В" in message


def test_import_deduplicates_repeated_predecessor():
    data = make_xlsx([
        ["Проектирование", "", "", 5, ""],
        ["Разработка", "", "", 10, "Проектирование, Проектирование"],
    ])
    tasks, errors = excel_io.parse_excel(data)
    assert errors == []
    assert next(t for t in tasks if t["name"] == "Разработка")["predecessors"] == [1]


def test_import_rejects_fractional_duration():
    data = make_xlsx([["Разработка", "", "", 5.7, ""]])
    tasks, errors = excel_io.parse_excel(data)
    assert tasks == []
    assert errors[0]["column"] == "Длительность"


def test_import_accepts_whole_float_duration():
    data = make_xlsx([["Разработка", "", "", 5.0, ""]])
    tasks, errors = excel_io.parse_excel(data)
    assert errors == []
    assert tasks[0]["duration"] == 5


def test_import_rejects_a_file_that_is_not_a_workbook():
    tasks, errors = excel_io.parse_excel(b"not an xlsx file")
    assert tasks == []
    assert len(errors) == 1
    assert errors[0]["row"] == 0


def test_import_rejects_negative_duration():
    data = make_xlsx([["Разработка", "", "", -3, ""]])
    tasks, errors = excel_io.parse_excel(data)
    assert tasks == []
    assert errors[0]["column"] == "Длительность"


def test_import_rejects_a_file_with_missing_column():
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Задача", "Описание", "Исполнитель", "Длительность"])  # нет предшественников
    sheet.append(["Разработка", "", "", 5])
    buffer = io.BytesIO()
    workbook.save(buffer)
    tasks, errors = excel_io.parse_excel(buffer.getvalue())
    assert tasks == []
    assert errors[0]["row"] == 1
    assert "Предшественники" in errors[0]["message"]


def test_import_reports_a_row_with_no_name_but_other_data():
    data = make_xlsx([["", "", "", 0, ""]])
    tasks, errors = excel_io.parse_excel(data)
    assert tasks == []
    assert errors[0]["column"] == "Задача"


def test_export_writes_predecessor_names_not_ids():
    set_tasks(branching_plan())
    data = excel_io.export_excel(srv.recalculate())
    from openpyxl import load_workbook
    sheet = load_workbook(io.BytesIO(data)).active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    testing = next(r for r in rows if r[0] == "Тестирование")
    assert testing[4] == "Разработка, Документация"


def test_export_then_import_gives_the_same_plan():
    set_tasks(branching_plan())
    original = srv.recalculate()
    data = excel_io.export_excel(original)
    tasks, errors = excel_io.parse_excel(data)
    assert errors == []
    srv._load_plan(tasks=tasks, project_start=original["project_start"])
    restored = srv.recalculate()
    strip = lambda plan: [(t["name"], t["duration"], t["start"], t["finish"]) for t in plan["tasks"]]
    assert strip(restored) == strip(original)


def test_export_import_keeps_the_constraint():
    set_tasks(branching_plan())
    srv.set_constraint(task_id=2, not_earlier_than="2026-09-20")
    original = srv.recalculate()
    tasks, errors = excel_io.parse_excel(excel_io.export_excel(original))
    assert errors == []
    srv._load_plan(tasks=tasks, project_start=original["project_start"])
    assert by_name(srv.recalculate())["Разработка"]["start"] == "2026-09-20"


def test_add_task_rejects_comma_in_name():
    set_tasks(branching_plan())
    before = srv.recalculate()
    try:
        srv.add_task(name="Разработка, часть 2", duration=3)
    except ValueError as e:
        assert "запятую" in str(e)
    else:
        raise AssertionError("запятая в имени должна была вызвать ValueError")
    assert srv.recalculate() == before


def test_import_rejects_comma_in_name():
    data = make_xlsx([["Разработка, часть 2", "", "", 5, ""]])
    tasks, errors = excel_io.parse_excel(data)
    assert tasks == []
    assert errors[0]["column"] == "Задача"


def test_import_ignores_a_foreign_sixth_column():
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(excel_io.COLUMNS + ["Комментарий"])
    sheet.append(["Разработка", "", "", 5, "", "2026-09-20"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    tasks, errors = excel_io.parse_excel(buffer.getvalue())
    assert errors == []
    assert tasks[0]["not_earlier_than"] is None


def test_import_accepts_a_real_date_cell_in_the_constraint_column():
    from datetime import date as date_type
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(excel_io.COLUMNS + [excel_io.CONSTRAINT_COLUMN])
    sheet.append(["Разработка", "", "", 5, "", date_type(2026, 9, 20)])
    buffer = io.BytesIO()
    workbook.save(buffer)
    tasks, errors = excel_io.parse_excel(buffer.getvalue())
    assert errors == []
    assert tasks[0]["not_earlier_than"] == "2026-09-20"


def test_add_task_rejects_duplicate_name():
    set_tasks(branching_plan())
    before = srv.recalculate()
    try:
        srv.add_task(name="Разработка", duration=2)
    except ValueError as e:
        assert "Разработка" in str(e)
    else:
        raise AssertionError("повторяющееся имя должно было вызвать ValueError")
    assert srv.recalculate() == before


def test_export_refuses_duplicate_names():
    set_tasks([task(1, "Тест", 2), task(2, "Тест", 3)])
    try:
        excel_io.export_excel(srv.recalculate())
    except ValueError as e:
        assert "Тест" in str(e)
    else:
        raise AssertionError("повторяющееся имя должно было вызвать ValueError")


def test_export_then_import_keeps_text_fields():
    set_tasks(branching_plan())
    srv.STATE["tasks"][2]["description"] = "Ядро и API"
    srv.STATE["tasks"][2]["assignee"] = "Иванов"
    original = srv.recalculate()
    tasks, errors = excel_io.parse_excel(excel_io.export_excel(original))
    assert errors == []
    srv._load_plan(tasks=tasks, project_start=original["project_start"])
    restored = by_name(srv.recalculate())["Разработка"]
    assert restored["description"] == "Ядро и API"
    assert restored["assignee"] == "Иванов"


class FakeTool:
    def __init__(self, name):
        self.name = name
        self.description = "описание"
        self.input_schema = {"type": "object", "properties": {}}


class FakeToolsResult:
    def __init__(self, names):
        self.tools = [FakeTool(n) for n in names]


class FakeSession:
    """Подменяет MCP-сессию: список инструментов и вызовы без реального сервера."""

    def __init__(self, names):
        self._names = names
        self.calls = []

    async def list_tools(self):
        return FakeToolsResult(self._names)

    async def call_tool(self, name, arguments):
        self.calls.append((name, arguments))

        class Block:
            text = "готово"

        class Result:
            content = [Block()]
            is_error = False

        return Result()


def test_service_tools_are_hidden_from_the_model():
    import asyncio
    session = FakeSession(["get_plan", "shift_tasks", "_load_plan"])
    tools = asyncio.run(agent_loop.tools_for_model(session))
    names = [t["function"]["name"] for t in tools]
    assert names == ["get_plan", "shift_tasks"]
    assert tools[0]["type"] == "function"
    assert tools[0]["function"]["parameters"] == {"type": "object", "properties": {}}


def test_broken_tool_arguments_do_not_crash_the_loop():
    arguments, error = agent_loop.parse_arguments('{"task_ids": [1,')
    assert arguments is None
    assert "JSON" in error


def test_good_tool_arguments_parse():
    arguments, error = agent_loop.parse_arguments('{"days": 7}')
    assert arguments == {"days": 7}
    assert error is None


def test_call_tool_turns_a_raised_error_into_text():
    import asyncio

    class RaisingSession:
        async def call_tool(self, name, arguments):
            raise RuntimeError("нет такого инструмента")

    text = asyncio.run(agent_loop.call_tool(RaisingSession(), "выдумка", {}))
    assert "нет такого инструмента" in text


def run_loop(session, replies, message="сдвинь задачи"):
    """Прогнать run_agent на поддельном транспорте, без сети."""
    import asyncio

    def handler(request):
        return httpx.Response(200, json=replies.pop(0))

    os.environ.setdefault("OPENROUTER_API_KEY", "test-key")
    return asyncio.run(agent_loop.run_agent(
        session, [{"role": "user", "content": message}],
        transport=httpx.MockTransport(handler)))


def assistant_asking(tool_call_id, name, arguments):
    return {"choices": [{"message": {"role": "assistant", "content": None,
            "tool_calls": [{"id": tool_call_id, "type": "function",
                            "function": {"name": name, "arguments": arguments}}]}}]}


def assistant_saying(text):
    return {"choices": [{"message": {"role": "assistant", "content": text}}]}


def test_run_agent_executes_the_tool_then_answers():
    session = FakeSession(["get_plan", "shift_tasks", "_load_plan"])
    text, history, gave_up = run_loop(session, [
        assistant_asking("c1", "shift_tasks", '{"task_ids": [1], "days": 7}'),
        assistant_saying("Сдвинул на неделю."),
    ])
    assert text == "Сдвинул на неделю."
    assert gave_up is False
    assert session.calls == [("shift_tasks", {"task_ids": [1], "days": 7})]
    assert any(m.get("role") == "tool" and m["tool_call_id"] == "c1" for m in history)


def test_run_agent_refuses_a_tool_the_model_was_never_shown():
    session = FakeSession(["get_plan", "shift_tasks", "_load_plan"])
    text, history, gave_up = run_loop(session, [
        assistant_asking("c1", "_load_plan", '{"tasks": [], "project_start": "2026-09-01"}'),
        assistant_saying("Не могу."),
    ])
    assert session.calls == []  # служебный инструмент не выполнен
    answer = next(m for m in history if m.get("role") == "tool")
    assert "недоступен" in answer["content"]


def test_run_agent_stops_at_the_ceiling():
    session = FakeSession(["get_plan", "shift_tasks"])
    replies = [assistant_asking(f"c{i}", "get_plan", "{}") for i in range(agent_loop.MAX_STEPS)]
    text, history, gave_up = run_loop(session, replies)
    assert gave_up is True
    assert len(session.calls) == agent_loop.MAX_STEPS
    assert text


class InProcessMcp:
    """Подменяет MCP-сессию: зовёт инструменты напрямую, без подпроцесса."""

    async def call_tool(self, name, arguments):
        text = json.dumps(getattr(srv, name)(**arguments), ensure_ascii=False)

        class Block:
            pass

        block = Block()
        block.text = text

        class Result:
            content = [block]
            is_error = False

        return Result()


def test_chat_rolls_back_and_forgets_the_phrase_when_the_model_fails():
    import asyncio

    import api
    from fastapi import HTTPException

    set_tasks(branching_plan())
    api.app.state.mcp = InProcessMcp()
    api.history.clear()
    before = srv.recalculate()

    async def failing_run(session, messages, transport=None):
        srv.shift_tasks(task_ids=[1], days=5)  # часть работы уже применена
        raise RuntimeError("модель недоступна")

    original = agent_loop.run_agent
    agent_loop.run_agent = failing_run
    try:
        asyncio.run(api.chat(api.ChatRequest(message="сдвинь всё")))
    except HTTPException as error:
        assert error.status_code == 502
    else:
        raise AssertionError("отказ модели должен был дать 502")
    finally:
        agent_loop.run_agent = original

    assert srv.recalculate() == before  # частичная правка отменена
    assert api.history == []            # фраза не осталась висеть


def test_deadline_is_absent_until_it_is_set():
    set_tasks(branching_plan())
    plan = srv.recalculate()
    assert plan["deadline"] is None
    assert plan["fits_deadline"] is None
    assert plan["days_late"] == 0


def test_deadline_counts_the_named_day_as_the_last_working_one():
    set_tasks(branching_plan())  # проект заканчивается 2026-09-17
    fits = srv.set_deadline(deadline="2026-09-17")
    assert fits["fits_deadline"] is True  # финиш 17-го, успеть «к концу 17-го» — да
    assert fits["days_late"] == 0
    late = srv.set_deadline(deadline="2026-09-14")
    assert late["fits_deadline"] is False
    assert late["days_late"] == 2  # граница — конец 14-го, то есть 15-е


def test_deadline_survives_a_change_and_shows_the_delay():
    set_tasks(branching_plan())
    srv.set_deadline(deadline="2026-09-17")
    plan = srv.set_duration(task_id=4, duration=9)  # «Тестирование» на 5 дней дольше
    assert plan["deadline"] == "2026-09-17"
    assert plan["fits_deadline"] is False
    # Финиш уехал с 17-го на 22-е, а граница — конец 17-го, то есть 18-е.
    # Один день из пяти съел запас, который был до дедлайна.
    assert plan["days_late"] == 4


def test_deadline_can_be_cleared():
    set_tasks(branching_plan())
    srv.set_deadline(deadline="2026-09-14")
    plan = srv.set_deadline(deadline=None)
    assert plan["deadline"] is None
    assert plan["fits_deadline"] is None


def test_rollback_restores_the_deadline():
    set_tasks(branching_plan())
    srv.set_deadline(deadline="2026-09-17")
    snapshot = srv.recalculate()
    srv.set_deadline(deadline="2026-10-01")
    srv._load_plan(tasks=snapshot["tasks"], project_start=snapshot["project_start"],
                   deadline=snapshot["deadline"])
    assert srv.recalculate()["deadline"] == "2026-09-17"


def test_sample_data_loads_and_has_a_critical_path():
    """Этими данными наполняется план при старте приложения."""
    import sample_data

    plan = srv._load_plan(tasks=sample_data.sample_tasks(), project_start="2026-09-01")
    assert len(plan["tasks"]) == 10
    assert any(t["on_critical_path"] for t in plan["tasks"])
    assert any(not t["on_critical_path"] for t in plan["tasks"])  # есть и запас, и критический путь
    assert all(t["predecessors"] or t["name"] == "Сбор требований" for t in plan["tasks"])


if __name__ == "__main__":
    for name, fn in sorted(globals().items()):
        if name.startswith("test_"):
            fn()
            print("ok", name)
